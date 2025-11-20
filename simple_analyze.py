#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
简化版结果分析 - 避免编码问题
"""

import os
import sys
import json
from pathlib import Path
from datetime import datetime

# 强制使用UTF-8
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

def read_potfile_and_mapping():
    """读取potfile和映射文件"""
    output_dir = Path("batch_crack_output")
    potfile_path = output_dir / "batch_results.potfile"
    mapping_file = output_dir / "uuid_hash_mapping.json"

    if not potfile_path.exists():
        print(f"错误: Potfile不存在: {potfile_path}")
        return None, None

    if not mapping_file.exists():
        print(f"错误: UUID映射文件不存在: {mapping_file}")
        return None, None

    # 读取UUID映射
    print(f"读取UUID映射: {mapping_file}")
    with open(mapping_file, 'r', encoding='utf-8') as f:
        mapping_data = json.load(f)

    # 创建hash到UUID的反向映射
    hash_to_uuid = {}
    for hash_index, info in mapping_data.items():
        hash_to_uuid[info['hash']] = info['uuid']

    # 读取potfile结果
    results = {}
    print(f"读取破解结果: {potfile_path}")

    with open(potfile_path, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            line = line.strip()
            if ':' in line:
                # 格式: $jksprivk$...:password
                hash_part, password = line.rsplit(':', 1)

                # 通过hash找到对应的UUID
                if hash_part in hash_to_uuid:
                    uuid = hash_to_uuid[hash_part]
                    results[uuid] = password

    print(f"发现 {len(results)} 个破解成功的密码")
    return results, hash_to_uuid

def extract_keystore_info(keystore_path, password):
    """提取keystore信息"""
    import subprocess

    try:
        # 使用keytool获取信息
        cmd = [
            'keytool', '-list', '-v',
            '-keystore', str(keystore_path),
            '-storepass', password
        ]

        result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)

        if result.returncode != 0:
            return None, None, None, 'JKS'

        output = result.stdout

        # 提取别名
        alias = None
        for line in output.split('\n'):
            if 'Alias name:' in line:
                alias = line.split(':', 1)[1].strip()
                break

        # 提取MD5和SHA1指纹
        md5_fingerprint = None
        sha1_fingerprint = None

        for line in output.split('\n'):
            line_lower = line.lower()
            if 'md5' in line_lower and ':' in line:
                # 提取MD5值并清理
                md5_part = line.split(':', 1)[1].strip()
                md5_fingerprint = md5_part.replace(':', '').replace(' ', '').upper()
            elif 'sha1' in line_lower and ':' in line:
                # 提取SHA1值并清理
                sha1_part = line.split(':', 1)[1].strip()
                sha1_fingerprint = sha1_part.replace(':', '').replace(' ', '').upper()

        # 确定keystore类型
        keystore_type = 'PKCS12' if 'PKCS12' in output else 'JKS'

        return alias, md5_fingerprint, sha1_fingerprint, keystore_type

    except Exception as e:
        print(f"警告: 提取信息失败 {keystore_path}: {e}")
        return None, None, None, 'JKS'

def process_results(cracked_passwords):
    """处理破解结果"""
    certificate_dir = Path("certificate")
    complete_results = []

    # 映射UUID到keystore路径
    keystore_map = {}
    for uuid_dir in certificate_dir.iterdir():
        if uuid_dir.is_dir():
            keystore_file = uuid_dir / "apk.keystore"
            if keystore_file.exists():
                keystore_map[uuid_dir.name] = keystore_file

    print(f"\n开始提取完整信息 (共{len(cracked_passwords)}个)...")
    count = 0

    for uuid, password in cracked_passwords.items():
        count += 1
        if count % 1000 == 0:
            print(f"  进度: {count}/{len(cracked_passwords)}")

        if uuid in keystore_map:
            keystore_path = keystore_map[uuid]

            # 提取信息
            alias, md5, sha1, ks_type = extract_keystore_info(keystore_path, password)

            complete_results.append({
                'uuid': uuid,
                'keystore_path': str(keystore_path),
                'password': password,
                'alias': alias or '提取失败',
                'public_key_md5': md5 or '提取失败',
                'public_key_sha1': sha1 or '提取失败',
                'keystore_type': ks_type,
                'file_size': keystore_path.stat().st_size
            })

    return complete_results

def save_json_report(results):
    """保存JSON报告"""
    output_dir = Path("batch_crack_output")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    report = {
        "timestamp": datetime.now().isoformat(),
        "summary": {
            "total_cracked": len(results),
            "success_rate": "100%"
        },
        "results": results
    }

    json_file = output_dir / f"batch_crack_results_{timestamp}.json"
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, ensure_ascii=False)

    print(f"\nJSON报告已保存: {json_file}")
    return str(json_file)

def save_excel_report(results):
    """保存Excel报告"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("警告: 未安装openpyxl，跳过Excel报告生成")
        return None

    output_dir = Path("batch_crack_output")

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "批量破解结果"

    # 表头
    headers = [
        "UUID", "路径", "文件名", "别名", "私钥密码",
        "签名公钥MD5", "签名公钥SHA1", "keystore类型", "文件大小(字节)"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 填充数据
    for row, result in enumerate(results, 2):
        keystore_path = Path(result['keystore_path'])

        ws.cell(row=row, column=1, value=result['uuid'])
        ws.cell(row=row, column=2, value=str(keystore_path.parent))
        ws.cell(row=row, column=3, value=keystore_path.name)
        ws.cell(row=row, column=4, value=result['alias'])
        ws.cell(row=row, column=5, value=result['password'])
        ws.cell(row=row, column=6, value=result['public_key_md5'])
        ws.cell(row=row, column=7, value=result['public_key_sha1'])
        ws.cell(row=row, column=8, value=result['keystore_type'])
        ws.cell(row=row, column=9, value=result['file_size'])

    # 保存文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = output_dir / f"batch_crack_results_{timestamp}.xlsx"
    wb.save(excel_file)

    print(f"Excel报告已保存: {excel_file}")
    return str(excel_file)

def main():
    print("=" * 60)
    print("批量破解结果分析器 (简化版)")
    print("=" * 60)

    # 读取potfile和映射
    cracked_passwords, hash_to_uuid = read_potfile_and_mapping()

    if cracked_passwords is None:
        return 1

    if len(cracked_passwords) == 0:
        print("没有找到破解结果")
        return 1

    # 处理结果
    complete_results = process_results(cracked_passwords)

    # 保存报告
    save_json_report(complete_results)
    save_excel_report(complete_results)

    print("\n" + "=" * 60)
    print(f"分析完成! 共处理 {len(complete_results)} 个keystore")
    print("=" * 60)

    return 0

if __name__ == "__main__":
    sys.exit(main())
