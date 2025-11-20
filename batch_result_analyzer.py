#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
批量破解结果分析器
将hashcat的potfile结果映射回原始keystore文件并提取完整信息
"""

import os
import json
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn

# 导入现有的信息提取器
try:
    from keystore_info_extractor import KeystoreInfoExtractor
except ImportError:
    print("错误: 无法导入keystore_info_extractor模块")
    import sys
    sys.exit(1)

console = Console()

class BatchResultAnalyzer:
    def __init__(self):
        self.certificate_dir = Path("certificate")
        self.output_dir = Path("batch_crack_output")
        self.potfile_path = self.output_dir / "batch_results.potfile"
        self.mapping_file = self.output_dir / "uuid_hash_mapping.json"
        self.keystore_extractor = KeystoreInfoExtractor()
        
        # 结果统计
        self.stats = {
            'total_keystores': 0,
            'cracked_passwords': 0,
            'failed_info_extraction': 0,
            'successful_complete_info': 0
        }
    
    def read_potfile_results(self) -> Dict[str, str]:
        """读取hashcat potfile结果并通过映射文件匹配UUID"""
        if not self.potfile_path.exists():
            console.print(f"[red]❌ Potfile不存在: {self.potfile_path}[/red]")
            return {}
        
        if not self.mapping_file.exists():
            console.print(f"[red]❌ UUID映射文件不存在: {self.mapping_file}[/red]")
            return {}
        
        # 读取UUID映射
        console.print(f"[cyan]📖 读取UUID映射: {self.mapping_file}[/cyan]")
        with open(self.mapping_file, 'r', encoding='utf-8') as f:
            mapping_data = json.load(f)
        
        # 创建hash到UUID的反向映射
        hash_to_uuid = {}
        for hash_index, info in mapping_data.items():
            hash_to_uuid[info['hash']] = info['uuid']
        
        # 读取potfile结果
        results = {}
        console.print(f"[cyan]📖 读取破解结果: {self.potfile_path}[/cyan]")
        
        with open(self.potfile_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if ':' in line:
                    # 格式: $jksprivk$...:password
                    hash_part, password = line.rsplit(':', 1)
                    
                    # 通过hash找到对应的UUID
                    if hash_part in hash_to_uuid:
                        uuid = hash_to_uuid[hash_part]
                        results[uuid] = password
                    else:
                        console.print(f"[yellow]⚠️ 找不到hash对应的UUID: {hash_part[:50]}...[/yellow]")
        
        console.print(f"[green]✅ 发现 {len(results)} 个破解成功的密码[/green]")
        return results
    
    def map_keystores(self) -> Dict[str, Path]:
        """映射UUID到keystore文件路径"""
        keystore_map = {}
        
        for uuid_dir in self.certificate_dir.iterdir():
            if uuid_dir.is_dir():
                keystore_file = uuid_dir / "apk.keystore"
                if keystore_file.exists():
                    keystore_map[uuid_dir.name] = keystore_file
        
        return keystore_map
    
    def extract_complete_info(self, uuid: str, keystore_path: Path, password: str) -> Dict:
        """提取keystore的完整信息"""
        try:
            # 使用现有的信息提取器
            alias, public_key_md5, public_key_sha1, keystore_type = self.keystore_extractor.extract_simple_info(
                str(keystore_path), password
            )
            
            return {
                'uuid': uuid,
                'keystore_path': str(keystore_path),
                'password': password,
                'alias': alias,
                'public_key_md5': public_key_md5,
                'public_key_sha1': public_key_sha1,
                'keystore_type': keystore_type,
                'file_size': keystore_path.stat().st_size,
                'extraction_success': True,
                'extraction_error': None
            }
        except Exception as e:
            return {
                'uuid': uuid,
                'keystore_path': str(keystore_path),
                'password': password,
                'alias': '提取失败',
                'public_key_md5': '提取失败',
                'public_key_sha1': '提取失败',
                'keystore_type': 'JKS',
                'file_size': keystore_path.stat().st_size,
                'extraction_success': False,
                'extraction_error': str(e)
            }
    
    def process_all_results(self, cracked_passwords: Dict[str, str], keystore_map: Dict[str, Path]) -> List[Dict]:
        """处理所有破解结果"""
        complete_results = []
        
        console.print(f"[cyan]🔍 提取完整信息...[/cyan]")
        
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            console=console
        ) as progress:
            
            task = progress.add_task("提取信息...", total=len(cracked_passwords))
            
            for uuid, password in cracked_passwords.items():
                if uuid in keystore_map:
                    keystore_path = keystore_map[uuid]
                    result = self.extract_complete_info(uuid, keystore_path, password)
                    complete_results.append(result)
                    
                    if result['extraction_success']:
                        self.stats['successful_complete_info'] += 1
                    else:
                        self.stats['failed_info_extraction'] += 1
                else:
                    console.print(f"[yellow]⚠️ 找不到UUID对应的keystore: {uuid}[/yellow]")
                
                progress.advance(task, 1)
        
        return complete_results
    
    def generate_excel_report(self, results: List[Dict]) -> str:
        """生成Excel报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            console.print("[yellow]⚠️ 未安装openpyxl，跳过Excel报告生成[/yellow]")
            return ""
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "批量破解结果"
        
        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 表头
        headers = [
            "UUID", "路径", "文件名", "别名", "私钥密码", 
            "签名公钥MD5", "签名公钥SHA1", "keystore类型", "文件大小(字节)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # 填充数据
        for row, result in enumerate(results, 2):
            keystore_path = Path(result['keystore_path'])
            
            ws.cell(row=row, column=1, value=result['uuid']).border = border
            ws.cell(row=row, column=2, value=str(keystore_path.parent)).border = border
            ws.cell(row=row, column=3, value=keystore_path.name).border = border
            ws.cell(row=row, column=4, value=result['alias']).border = border
            ws.cell(row=row, column=5, value=result['password']).border = border
            ws.cell(row=row, column=6, value=result['public_key_md5']).border = border
            ws.cell(row=row, column=7, value=result['public_key_sha1']).border = border
            ws.cell(row=row, column=8, value=result['keystore_type']).border = border
            ws.cell(row=row, column=9, value=result['file_size']).border = border
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = self.output_dir / f"batch_crack_results_{timestamp}.xlsx"
        wb.save(excel_file)
        
        return str(excel_file)
    
    def generate_json_report(self, results: List[Dict]) -> str:
        """生成JSON报告"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        report = {
            "timestamp": datetime.now().isoformat(),
            "summary": {
                "total_cracked": len(results),
                "successful_info_extraction": self.stats['successful_complete_info'],
                "failed_info_extraction": self.stats['failed_info_extraction'],
                "success_rate": f"{(self.stats['successful_complete_info']/max(len(results),1)*100):.1f}%"
            },
            "results": results
        }
        
        json_file = self.output_dir / f"batch_crack_results_{timestamp}.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, ensure_ascii=False)
        
        return str(json_file)
    
    def show_summary_table(self, results: List[Dict]):
        """显示结果汇总表"""
        if not results:
            console.print("[yellow]⚠️ 没有破解结果可显示[/yellow]")
            return
        
        # 显示前10个结果
        table = Table(title="🎉 批量破解结果汇总 (前10个)", border_style="green")
        table.add_column("UUID", style="cyan", width=12)
        table.add_column("密码", style="yellow", width=10)
        table.add_column("别名", style="blue", width=15)
        table.add_column("MD5", style="green", width=12)
        table.add_column("SHA1", style="magenta", width=12)
        
        for result in results[:10]:
            table.add_row(
                result['uuid'][:12] + "...",
                result['password'],
                result['alias'][:15] if len(result['alias']) > 15 else result['alias'],
                result['public_key_md5'][:12] + "..." if len(result['public_key_md5']) > 12 else result['public_key_md5'],
                result['public_key_sha1'][:12] + "..." if len(result['public_key_sha1']) > 12 else result['public_key_sha1']
            )
        
        console.print(table)
        
        if len(results) > 10:
            console.print(f"[dim]... 及其他 {len(results)-10} 个结果[/dim]")
    
    def analyze_and_report(self):
        """分析破解结果并生成报告"""
        console.print(Panel.fit(
            "[bold cyan]🔍 批量破解结果分析器[/bold cyan]\n"
            "分析potfile结果并生成完整报告",
            border_style="cyan"
        ))
        
        # 1. 读取potfile结果
        cracked_passwords = self.read_potfile_results()
        if not cracked_passwords:
            console.print("[red]❌ 没有找到破解结果[/red]")
            return False
        
        self.stats['cracked_passwords'] = len(cracked_passwords)
        
        # 2. 映射keystore文件
        keystore_map = self.map_keystores()
        self.stats['total_keystores'] = len(keystore_map)
        
        # 3. 提取完整信息
        complete_results = self.process_all_results(cracked_passwords, keystore_map)
        
        if not complete_results:
            console.print("[red]❌ 无法处理任何结果[/red]")
            return False
        
        # 4. 显示汇总表
        self.show_summary_table(complete_results)
        
        # 5. 生成报告
        console.print("\n[cyan]📊 生成详细报告...[/cyan]")
        
        json_file = self.generate_json_report(complete_results)
        excel_file = self.generate_excel_report(complete_results)
        
        # 6. 显示统计信息
        stats_table = Table(title="📈 处理统计", border_style="blue")
        stats_table.add_column("项目", style="cyan")
        stats_table.add_column("数值", style="white")
        
        stats_table.add_row("破解成功密码", str(self.stats['cracked_passwords']))
        stats_table.add_row("信息提取成功", str(self.stats['successful_complete_info']))
        stats_table.add_row("信息提取失败", str(self.stats['failed_info_extraction']))
        stats_table.add_row("整体成功率", f"{(self.stats['successful_complete_info']/max(self.stats['cracked_passwords'],1)*100):.1f}%")
        
        console.print(stats_table)
        
        # 7. 显示文件位置
        console.print("\n[bold green]📁 生成的报告文件:[/bold green]")
        console.print(f"[cyan]JSON报告:[/cyan] {json_file}")
        if excel_file:
            console.print(f"[cyan]Excel报告:[/cyan] {excel_file}")
        
        console.print("\n[bold yellow]💡 报告包含的信息:[/bold yellow]")
        console.print("[yellow]- UUID文件夹名作为唯一标识[/yellow]")
        console.print("[yellow]- 破解的6位密码[/yellow]")
        console.print("[yellow]- keystore别名信息[/yellow]")
        console.print("[yellow]- 公钥MD5和SHA1哈希值（无冒号格式）[/yellow]")
        console.print("[yellow]- keystore类型和文件信息[/yellow]")
        
        return True

def main():
    analyzer = BatchResultAnalyzer()
    success = analyzer.analyze_and_report()
    return 0 if success else 1

if __name__ == "__main__":
    import sys
    sys.exit(main()) 