#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""破解进度管理器

管理批量keystore破解任务的会话状态、断点续传、进度跟踪、结果导出，
使用MD5生成会话ID，JSON格式保存进度文件，10秒自动保存，
支持rich.prompt交互式恢复确认，导出Excel/JSON格式化结果报告。

Architecture:
    会话创建 → 任务跟踪 → 自动保存 → 断点续传 → 结果导出

    ProgressManager (progress_manager.py:77)
        ├─ __init__() (L80): 初始化progress目录 + 10秒自动保存间隔
        ├─ generate_session_id() (L87): MD5(target+mask+mode+date)[:12]生成会话ID
        ├─ create_session() (L92): 创建新会话或检测并恢复未完成会话
        ├─ _confirm_resume() (L137): rich.prompt.Confirm交互式确认恢复
        ├─ get_pending_tasks() (L146): 过滤status="pending"的任务列表
        ├─ get_task_by_path() (L154): 通过file_path查找TaskProgress
        ├─ start_task() (L164): 更新status="processing" + start_time + attempts++
        ├─ complete_task() (L177): 更新status="completed" + password + 证书信息（7个字段）
        ├─ fail_task() (L202): 更新status="failed" + error_message
        ├─ skip_task() (L218): 更新status="skipped" + reason
        ├─ _update_estimated_completion() (L234): 基于平均duration计算预估完成时间
        ├─ _auto_save() (L256): 10秒间隔自动保存progress/{session_id}.json
        ├─ save_session() (L263): JSON序列化BatchProgress到文件
        ├─ load_session() (L276): JSON反序列化恢复BatchProgress对象
        ├─ list_sessions() (L291): glob扫描progress/*.json返回session_id列表
        ├─ delete_session() (L296): 删除进度文件
        ├─ show_progress() (L308): rich Table显示进度（9行：目标/掩码/模式/总数/已完成/已失败/已跳过/进度%/预计完成）
        ├─ get_results_summary() (L340): 生成结果汇总字典（10个统计字段 + 破解结果列表）
        ├─ export_results() (L379): 导出JSON和XLSX（2个工作表）
        ├─ _export_to_xlsx() (L424): openpyxl生成Excel（破解结果表 + 统计信息表）
        └─ cleanup_completed_sessions() (L515): 清理7天前已完成的旧会话

    TaskProgress (dataclass, progress_manager.py:21)
        ├─ 11个字段：file_path, status, password, start_time, end_time, duration,
        │            error_message, attempts, alias, public_key_md5, public_key_sha1,
        │            keystore_type, certificate_info
        ├─ to_dict() (L38): 使用asdict()转换为字典
        └─ from_dict() (L42): 类方法反序列化

    BatchProgress (dataclass, progress_manager.py:45)
        ├─ 10个字段：session_id, target_path, mask, mode, total_files,
        │            completed_files, failed_files, skipped_files, start_time,
        │            last_update, estimated_completion, tasks (List[TaskProgress])
        ├─ to_dict() (L65): 递归序列化tasks列表
        └─ from_dict() (L71): 类方法递归反序列化

Features:
    - 会话ID生成：MD5(target_path_mask_mode_YYYYMMDD)[:12] (progress_manager.py:89-90)
    - 断点续传：检测未完成会话 + rich.prompt.Confirm交互式确认 (progress_manager.py:98-107)
    - 5种任务状态：pending, processing, completed, failed, skipped (progress_manager.py:25)
    - 自动保存：10秒间隔检查 + JSON序列化 (progress_manager.py:84, 256-261)
    - 证书信息跟踪：alias, public_key_md5, public_key_sha1, keystore_type, certificate_info (progress_manager.py:32-36)
    - 预估完成时间：基于已完成任务平均duration计算 (progress_manager.py:234-254)
    - UUID作为ID：使用file_path.parent.name（UUID文件夹名）(progress_manager.py:353)
    - Excel双表导出：破解结果表（9列）+ 统计信息表（10行）(progress_manager.py:438-509)
    - 会话清理：删除7天前且已完成（completed+failed+skipped=total）的会话 (progress_manager.py:515-538)

Args (方法参数):
    ProgressManager.__init__(progress_dir: str = "progress"):
        初始化进度管理器，创建progress目录

    create_session(target_path: str, mask: str, mode: str, file_list: List[str]) -> str:
        创建新会话或恢复未完成会话，返回session_id

    complete_task(file_path: str, password: str, duration: float,
                  alias: str = None, public_key_md5: str = None, public_key_sha1: str = None,
                  keystore_type: str = None, certificate_info: Dict[str, Any] = None) -> bool:
        完成任务并记录证书信息（7个字段）

    export_results(output_file: Optional[str] = None, export_xlsx: bool = True) -> str:
        导出结果到JSON和XLSX，返回文件路径

        示例：
        # 初始化进度管理器
        pm = ProgressManager(progress_dir="progress")

        # 创建会话
        files = ["cert1/apk.keystore", "cert2/apk.keystore"]
        session_id = pm.create_session("/path/to/certs", "?a?a?a?a?a?a", "batch", files)

        # 处理任务
        for task in pm.get_pending_tasks():
            pm.start_task(task.file_path)
            # ... 破解逻辑 ...
            pm.complete_task(task.file_path, password="123456", duration=120.5,
                           alias="mykey", public_key_md5="A1B2C3...", public_key_sha1="D4E5F6...")

        # 导出结果
        xlsx_path = pm.export_results(export_xlsx=True)

Returns (返回值):
    TaskProgress对象（11个字段）:
        file_path (str): keystore文件路径
        status (str): pending/processing/completed/failed/skipped
        password (str): 破解的密码
        start_time (str): ISO8601开始时间
        end_time (str): ISO8601结束时间
        duration (float): 破解耗时（秒）
        error_message (str): 错误信息
        attempts (int): 尝试次数
        alias (str): keystore别名
        public_key_md5 (str): 公钥MD5哈希
        public_key_sha1 (str): 公钥SHA1哈希
        keystore_type (str): "JKS" or "PKCS12"
        certificate_info (Dict): 证书详细信息

    BatchProgress对象（10个字段 + tasks列表）:
        session_id (str): 12位MD5会话ID
        target_path (str): 目标路径
        mask (str): 密码掩码
        mode (str): 破解模式
        total_files (int): 总文件数
        completed_files (int): 已完成数
        failed_files (int): 已失败数
        skipped_files (int): 已跳过数
        start_time (str): ISO8601开始时间
        last_update (str): ISO8601最后更新时间
        estimated_completion (str): ISO8601预估完成时间
        tasks (List[TaskProgress]): 任务列表

    导出文件（2种格式）:
        {base_name}.json: 结果汇总JSON（10个统计字段 + 破解结果列表）
        {base_name}.xlsx: Excel报告（2个工作表：破解结果9列 + 统计信息10行）

Requirements:
    - rich (Console, Table, Panel, Confirm)
    - openpyxl (可选，Excel导出)
    - Python标准库: json, time, hashlib, pathlib, datetime, dataclasses

Technical Notes:
    会话ID生成策略:
        内容: f"{target_path}_{mask}_{mode}_{YYYYMMDD}" (progress_manager.py:89)
        哈希: hashlib.md5(content.encode()).hexdigest()[:12] (progress_manager.py:90)
        示例: "a1b2c3d4e5f6"（12位十六进制）

    断点续传流程:
        1. 检测未完成会话：load_session(session_id) (progress_manager.py:98)
        2. 显示会话信息：目标/进度 (progress_manager.py:100-102)
        3. 交互式确认：rich.prompt.Confirm.ask() (progress_manager.py:104, 141)
        4. 恢复或创建新会话 (progress_manager.py:105-109)

    任务状态机:
        pending → processing (start_task) → completed/failed/skipped
        状态字段: task.status (progress_manager.py:25)
        状态更新: L170, L185, L208, L224

    自动保存机制:
        间隔: 10秒 (progress_manager.py:84)
        检查: time.time() - last_save_time >= auto_save_interval (progress_manager.py:259)
        触发: start_task, complete_task, fail_task, skip_task调用_auto_save() (progress_manager.py:174, 199, 215, 231)

    预估完成时间计算:
        平均耗时: sum(task.duration) / len(completed_tasks) (progress_manager.py:246)
        剩余文件: total - completed - failed - skipped (progress_manager.py:247-250)
        预估秒数: remaining * avg_duration (progress_manager.py:252)
        完成时间: datetime.now().timestamp() + estimated_seconds (progress_manager.py:253-254)

    UUID作为ID设计:
        ID字段: file_path.parent.name (progress_manager.py:353)
        原因: certificate/[UUID]/apk.keystore结构，UUID保证唯一性
        用途: Excel导出的"ID"列，便于数据库关联

    Excel导出结构:
        工作表1 - 破解结果 (progress_manager.py:438-483):
            9列: 路径, ID, 文件名, 别名, 私钥密码, 签名公钥MD5, 签名公钥SHA1, keystore类型, 破解耗时
            样式: 蓝色表头（#366092） + 边框 + 自动列宽（最大50）

        工作表2 - 统计信息 (progress_manager.py:485-509):
            10行: 会话ID, 破解时间, 目标路径, 密码掩码, 破解模式, 总文件数, 成功破解, 破解失败, 跳过文件, 成功率
            样式: 加粗键名 + 边框

    会话清理策略:
        保留天数: 7天 (progress_manager.py:515)
        清理条件: 文件修改时间 < 当前时间-7天 AND 已完成（completed+failed+skipped>=total）(progress_manager.py:524-531)
        文件操作: session_file.unlink() (progress_manager.py:532)

Workflow:
    1. 初始化ProgressManager，创建progress目录
    2. 调用create_session(target_path, mask, mode, file_list)
    3. 生成会话ID：MD5(target_mask_mode_date)[:12]
    4. 检测未完成会话并交互式确认恢复
    5. 创建BatchProgress对象并初始化所有TaskProgress（status="pending"）
    6. 保存会话到progress/{session_id}.json
    7. 循环处理任务：
       - get_pending_tasks()获取待处理任务
       - start_task(file_path)更新status="processing"
       - 破解成功：complete_task()记录password和证书信息
       - 破解失败：fail_task()记录error_message
       - 自动保存：每10秒检查并保存
    8. 更新预估完成时间（基于平均duration）
    9. show_progress()显示rich Table进度
    10. 导出结果：
        - get_results_summary()生成汇总字典
        - export_results()导出JSON和Excel（2个工作表）
    11. cleanup_completed_sessions()清理7天前旧会话

Author: Forensic Keystore Cracker Project
Version: 1.0.0
License: 仅用于授权的数字取证和安全研究
"""

import os
import json
import time
import hashlib
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any
from dataclasses import dataclass, asdict
from rich.console import Console
from rich.panel import Panel
from rich.table import Table

console = Console()

@dataclass
class TaskProgress:
    """单个任务进度"""
    file_path: str
    status: str  # pending, processing, completed, failed, skipped
    password: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    duration: Optional[float] = None
    error_message: Optional[str] = None
    attempts: int = 0
    alias: Optional[str] = None
    public_key_md5: Optional[str] = None
    public_key_sha1: Optional[str] = None
    keystore_type: Optional[str] = None
    certificate_info: Optional[Dict[str, Any]] = None
    
    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)
    
    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'TaskProgress':
        return cls(**data)

@dataclass
class BatchProgress:
    """批量任务进度"""
    session_id: str
    target_path: str
    mask: str
    mode: str
    total_files: int
    completed_files: int
    failed_files: int
    skipped_files: int
    start_time: str
    last_update: str
    estimated_completion: Optional[str] = None
    tasks: List[TaskProgress] = None
    
    def __post_init__(self):
        if self.tasks is None:
            self.tasks = []
    
    def to_dict(self) -> Dict[str, Any]:
        data = asdict(self)
        data['tasks'] = [task.to_dict() for task in self.tasks]
        return data
    
    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'BatchProgress':
        tasks_data = data.pop('tasks', [])
        progress = cls(**data)
        progress.tasks = [TaskProgress.from_dict(task) for task in tasks_data]
        return progress

class ProgressManager:
    """进度管理器"""
    
    def __init__(self, progress_dir: str = "progress"):
        self.progress_dir = Path(progress_dir)
        self.progress_dir.mkdir(exist_ok=True)
        self.current_session: Optional[BatchProgress] = None
        self.auto_save_interval = 10  # 秒
        self.last_save_time = 0
        
    def generate_session_id(self, target_path: str, mask: str, mode: str) -> str:
        """生成会话ID"""
        content = f"{target_path}_{mask}_{mode}_{datetime.now().strftime('%Y%m%d')}"
        return hashlib.md5(content.encode()).hexdigest()[:12]
    
    def create_session(self, target_path: str, mask: str, mode: str, 
                      file_list: List[str]) -> str:
        """创建新的破解会话"""
        session_id = self.generate_session_id(target_path, mask, mode)
        
        # 检查是否存在未完成的会话
        existing_session = self.load_session(session_id)
        if existing_session:
            console.print(f"[yellow]📄 发现未完成的会话: {session_id}[/yellow]")
            console.print(f"[cyan]   目标: {existing_session.target_path}[/cyan]")
            console.print(f"[cyan]   进度: {existing_session.completed_files}/{existing_session.total_files}[/cyan]")
            
            if self._confirm_resume():
                console.print("[green]✅ 恢复之前的会话[/green]")
                self.current_session = existing_session
                return session_id
            else:
                console.print("[yellow]🔄 创建新会话[/yellow]")
        
        # 创建新会话
        self.current_session = BatchProgress(
            session_id=session_id,
            target_path=target_path,
            mask=mask,
            mode=mode,
            total_files=len(file_list),
            completed_files=0,
            failed_files=0,
            skipped_files=0,
            start_time=datetime.now().isoformat(),
            last_update=datetime.now().isoformat()
        )
        
        # 初始化任务列表
        for file_path in file_list:
            task = TaskProgress(
                file_path=file_path,
                status="pending"
            )
            self.current_session.tasks.append(task)
        
        self.save_session()
        console.print(f"[green]✅ 创建新会话: {session_id}[/green]")
        return session_id
    
    def _confirm_resume(self) -> bool:
        """确认是否恢复会话"""
        try:
            from rich.prompt import Confirm
            return Confirm.ask("是否继续之前未完成的破解任务?")
        except:
            # 如果无法交互，默认恢复
            return True
    
    def get_pending_tasks(self) -> List[TaskProgress]:
        """获取待处理的任务"""
        if not self.current_session:
            return []
        
        return [task for task in self.current_session.tasks 
                if task.status == "pending"]
    
    def get_task_by_path(self, file_path: str) -> Optional[TaskProgress]:
        """根据文件路径获取任务"""
        if not self.current_session:
            return None
            
        for task in self.current_session.tasks:
            if task.file_path == file_path:
                return task
        return None
    
    def start_task(self, file_path: str) -> bool:
        """开始处理任务"""
        task = self.get_task_by_path(file_path)
        if not task:
            return False
            
        task.status = "processing"
        task.start_time = datetime.now().isoformat()
        task.attempts += 1
        
        self._auto_save()
        return True
    
    def complete_task(self, file_path: str, password: str, duration: float, 
                     alias: str = None, public_key_md5: str = None, public_key_sha1: str = None,
                     keystore_type: str = None, certificate_info: Dict[str, Any] = None) -> bool:
        """完成任务"""
        task = self.get_task_by_path(file_path)
        if not task:
            return False
            
        task.status = "completed"
        task.password = password
        task.end_time = datetime.now().isoformat()
        task.duration = duration
        task.alias = alias
        task.public_key_md5 = public_key_md5
        task.public_key_sha1 = public_key_sha1
        task.keystore_type = keystore_type
        task.certificate_info = certificate_info
        
        self.current_session.completed_files += 1
        self.current_session.last_update = datetime.now().isoformat()
        
        self._update_estimated_completion()
        self._auto_save()
        return True
    
    def fail_task(self, file_path: str, error_message: str) -> bool:
        """任务失败"""
        task = self.get_task_by_path(file_path)
        if not task:
            return False
            
        task.status = "failed"
        task.error_message = error_message
        task.end_time = datetime.now().isoformat()
        
        self.current_session.failed_files += 1
        self.current_session.last_update = datetime.now().isoformat()
        
        self._auto_save()
        return True
    
    def skip_task(self, file_path: str, reason: str) -> bool:
        """跳过任务"""
        task = self.get_task_by_path(file_path)
        if not task:
            return False
            
        task.status = "skipped"
        task.error_message = reason
        task.end_time = datetime.now().isoformat()
        
        self.current_session.skipped_files += 1
        self.current_session.last_update = datetime.now().isoformat()
        
        self._auto_save()
        return True
    
    def _update_estimated_completion(self):
        """更新预估完成时间"""
        if not self.current_session or self.current_session.completed_files == 0:
            return
            
        # 计算平均处理时间
        completed_tasks = [task for task in self.current_session.tasks 
                          if task.status == "completed" and task.duration]
        
        if not completed_tasks:
            return
            
        avg_duration = sum(task.duration for task in completed_tasks) / len(completed_tasks)
        remaining_files = (self.current_session.total_files - 
                          self.current_session.completed_files - 
                          self.current_session.failed_files - 
                          self.current_session.skipped_files)
        
        estimated_seconds = remaining_files * avg_duration
        estimated_time = datetime.now().timestamp() + estimated_seconds
        self.current_session.estimated_completion = datetime.fromtimestamp(estimated_time).isoformat()
    
    def _auto_save(self):
        """自动保存"""
        current_time = time.time()
        if current_time - self.last_save_time >= self.auto_save_interval:
            self.save_session()
            self.last_save_time = current_time
    
    def save_session(self):
        """保存会话"""
        if not self.current_session:
            return
            
        session_file = self.progress_dir / f"{self.current_session.session_id}.json"
        
        try:
            with open(session_file, 'w', encoding='utf-8') as f:
                json.dump(self.current_session.to_dict(), f, indent=2, ensure_ascii=False)
        except Exception as e:
            console.print(f"[red]❌ 保存会话失败: {e}[/red]")
    
    def load_session(self, session_id: str) -> Optional[BatchProgress]:
        """加载会话"""
        session_file = self.progress_dir / f"{session_id}.json"
        
        if not session_file.exists():
            return None
            
        try:
            with open(session_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return BatchProgress.from_dict(data)
        except Exception as e:
            console.print(f"[red]❌ 加载会话失败: {e}[/red]")
            return None
    
    def list_sessions(self) -> List[str]:
        """列出所有会话"""
        session_files = list(self.progress_dir.glob("*.json"))
        return [f.stem for f in session_files]
    
    def delete_session(self, session_id: str) -> bool:
        """删除会话"""
        session_file = self.progress_dir / f"{session_id}.json"
        
        try:
            if session_file.exists():
                session_file.unlink()
                return True
        except Exception as e:
            console.print(f"[red]❌ 删除会话失败: {e}[/red]")
        return False
    
    def show_progress(self):
        """显示当前进度"""
        if not self.current_session:
            console.print("[yellow]⚠️ 没有活动会话[/yellow]")
            return
            
        session = self.current_session
        
        # 计算进度百分比
        total_processed = session.completed_files + session.failed_files + session.skipped_files
        progress_percent = (total_processed / session.total_files * 100) if session.total_files > 0 else 0
        
        # 创建进度表格
        table = Table(title=f"📊 会话进度: {session.session_id}", border_style="blue")
        table.add_column("项目", style="cyan", width=15)
        table.add_column("值", style="white", width=30)
        
        table.add_row("目标路径", session.target_path)
        table.add_row("密码掩码", session.mask)
        table.add_row("破解模式", session.mode)
        table.add_row("总文件数", str(session.total_files))
        table.add_row("已完成", f"{session.completed_files} ✅")
        table.add_row("已失败", f"{session.failed_files} ❌")
        table.add_row("已跳过", f"{session.skipped_files} ⏭️")
        table.add_row("进度", f"{progress_percent:.1f}%")
        
        if session.estimated_completion:
            est_time = datetime.fromisoformat(session.estimated_completion)
            table.add_row("预计完成", est_time.strftime("%Y-%m-%d %H:%M:%S"))
        
        console.print(table)
    
    def get_results_summary(self) -> Dict[str, Any]:
        """获取结果汇总"""
        if not self.current_session:
            return {}
            
        successful_tasks = [task for task in self.current_session.tasks 
                           if task.status == "completed"]
        
        results = []
        for task in successful_tasks:
            file_path = Path(task.file_path)
            result_entry = {
                "路径": str(file_path.parent),
                "ID": file_path.parent.name,  # UUID文件夹名
                "文件名": file_path.name,
                "别名": task.alias or "未知",
                "私钥密码": task.password,
                "签名公钥MD5": task.public_key_md5 or "未提取",
                "签名公钥SHA1": task.public_key_sha1 or "未提取",
                "keystore类型": task.keystore_type or "JKS",
                "破解耗时": f"{task.duration:.2f}秒" if task.duration else "未知",
                "证书信息": task.certificate_info or {}
            }
            results.append(result_entry)
        
        return {
            "session_id": self.current_session.session_id,
            "破解时间": self.current_session.start_time,
            "目标路径": self.current_session.target_path,
            "密码掩码": self.current_session.mask,
            "破解模式": self.current_session.mode,
            "总文件数": self.current_session.total_files,
            "成功破解": len(successful_tasks),
            "破解失败": self.current_session.failed_files,
            "跳过文件": self.current_session.skipped_files,
            "成功率": f"{(len(successful_tasks) / max(self.current_session.total_files, 1) * 100):.1f}%",
            "破解结果": results
        }
    
    def export_results(self, output_file: Optional[str] = None, export_xlsx: bool = True) -> str:
        """导出结果到JSON和XLSX"""
        if not self.current_session:
            console.print("[red]❌ 没有活动会话可导出[/red]")
            return ""
            
        # 确定输出文件名
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = f"crack_results_{self.current_session.session_id}_{timestamp}"
            json_file = f"{base_name}.json"
            xlsx_file = f"{base_name}.xlsx"
        else:
            base_name = Path(output_file).stem
            json_file = f"{base_name}.json"
            xlsx_file = f"{base_name}.xlsx"
        
        results = self.get_results_summary()
        
        # 导出JSON文件
        try:
            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, ensure_ascii=False)
            console.print(f"[green]✅ JSON结果已导出到: {json_file}[/green]")
        except Exception as e:
            console.print(f"[red]❌ JSON导出失败: {e}[/red]")
            return ""
        
        # 导出XLSX文件
        if export_xlsx:
            try:
                xlsx_path = self._export_to_xlsx(results, xlsx_file)
                if xlsx_path:
                    console.print(f"[green]✅ Excel结果已导出到: {xlsx_path}[/green]")
                    return xlsx_path
            except Exception as e:
                console.print(f"[yellow]⚠️ Excel导出失败: {e}[/yellow]")
                console.print(f"[cyan]💡 提示: 请安装openpyxl: pip install openpyxl[/cyan]")
        
        # 如果没有导出Excel或Excel导出失败，但JSON成功，则给出提示
        if export_xlsx and not results.get("破解结果"):
            console.print(f"[yellow]ℹ️ 没有成功破解的结果，仅生成统计信息的Excel文件[/yellow]")
        
        return json_file
    
    def _export_to_xlsx(self, results: Dict[str, Any], xlsx_file: str) -> Optional[str]:
        """导出结果到Excel文件"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            console.print("[yellow]⚠️ 未安装openpyxl，跳过Excel导出[/yellow]")
            return None
        
        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        
        # 主结果表
        ws_main = wb.active
        ws_main.title = "破解结果"
        
        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 主表头
        headers = ["路径", "ID", "文件名", "别名", "私钥密码", "签名公钥MD5", "签名公钥SHA1", "keystore类型", "破解耗时"]
        for col, header in enumerate(headers, 1):
            cell = ws_main.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # 填充数据
        for row, result in enumerate(results.get("破解结果", []), 2):
            ws_main.cell(row=row, column=1, value=result.get("路径", "")).border = border
            ws_main.cell(row=row, column=2, value=result.get("ID", "")).border = border
            ws_main.cell(row=row, column=3, value=result.get("文件名", "")).border = border
            ws_main.cell(row=row, column=4, value=result.get("别名", "")).border = border
            ws_main.cell(row=row, column=5, value=result.get("私钥密码", "")).border = border
            ws_main.cell(row=row, column=6, value=result.get("签名公钥MD5", "")).border = border
            ws_main.cell(row=row, column=7, value=result.get("签名公钥SHA1", "")).border = border
            ws_main.cell(row=row, column=8, value=result.get("keystore类型", "")).border = border
            ws_main.cell(row=row, column=9, value=result.get("破解耗时", "")).border = border
        
        # 自动调整列宽
        for column in ws_main.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_main.column_dimensions[column_letter].width = adjusted_width
        
        # 统计信息表
        ws_stats = wb.create_sheet("统计信息")
        
        stats_data = [
            ["会话ID", results.get("session_id", "")],
            ["破解时间", results.get("破解时间", "")],
            ["目标路径", results.get("目标路径", "")],
            ["密码掩码", results.get("密码掩码", "")],
            ["破解模式", results.get("破解模式", "")],
            ["总文件数", results.get("总文件数", 0)],
            ["成功破解", results.get("成功破解", 0)],
            ["破解失败", results.get("破解失败", 0)],
            ["跳过文件", results.get("跳过文件", 0)],
            ["成功率", results.get("成功率", "0%")]
        ]
        
        for row, (key, value) in enumerate(stats_data, 1):
            ws_stats.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws_stats.cell(row=row, column=2, value=str(value))
            ws_stats.cell(row=row, column=1).border = border
            ws_stats.cell(row=row, column=2).border = border
        
        # 调整统计信息表列宽
        ws_stats.column_dimensions['A'].width = 15
        ws_stats.column_dimensions['B'].width = 30
        
        # 保存文件
        wb.save(xlsx_file)
        return xlsx_file
    
    def cleanup_completed_sessions(self, keep_days: int = 7):
        """清理已完成的旧会话"""
        current_time = time.time()
        cutoff_time = current_time - (keep_days * 24 * 60 * 60)
        
        cleaned = 0
        for session_file in self.progress_dir.glob("*.json"):
            try:
                # 检查文件修改时间
                if session_file.stat().st_mtime < cutoff_time:
                    session = self.load_session(session_file.stem)
                    if session:
                        # 只删除已完成的会话
                        total_processed = (session.completed_files + 
                                         session.failed_files + 
                                         session.skipped_files)
                        if total_processed >= session.total_files:
                            session_file.unlink()
                            cleaned += 1
            except Exception:
                continue
        
        if cleaned > 0:
            console.print(f"[green]🧹 清理了 {cleaned} 个已完成的旧会话[/green]") 