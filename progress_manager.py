#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
进度管理器 - 支持破解进度保存和断点续传
"""

import os
import json
import time
import hashlib
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any
from dataclasses import dataclass, asdict
from rich.console import Console
from rich.panel import Panel
from rich.table import Table

console = Console()

@dataclass
class TaskProgress:
    """单个任务进度"""
    file_path: str
    status: str  # pending, processing, completed, failed, skipped
    password: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    duration: Optional[float] = None
    error_message: Optional[str] = None
    attempts: int = 0
    alias: Optional[str] = None
    public_key_md5: Optional[str] = None
    public_key_sha1: Optional[str] = None
    keystore_type: Optional[str] = None
    certificate_info: Optional[Dict[str, Any]] = None
    
    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)
    
    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'TaskProgress':
        return cls(**data)

@dataclass
class BatchProgress:
    """批量任务进度"""
    session_id: str
    target_path: str
    mask: str
    mode: str
    total_files: int
    completed_files: int
    failed_files: int
    skipped_files: int
    start_time: str
    last_update: str
    estimated_completion: Optional[str] = None
    tasks: List[TaskProgress] = None
    
    def __post_init__(self):
        if self.tasks is None:
            self.tasks = []
    
    def to_dict(self) -> Dict[str, Any]:
        data = asdict(self)
        data['tasks'] = [task.to_dict() for task in self.tasks]
        return data
    
    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'BatchProgress':
        tasks_data = data.pop('tasks', [])
        progress = cls(**data)
        progress.tasks = [TaskProgress.from_dict(task) for task in tasks_data]
        return progress

class ProgressManager:
    """进度管理器"""
    
    def __init__(self, progress_dir: str = "progress"):
        self.progress_dir = Path(progress_dir)
        self.progress_dir.mkdir(exist_ok=True)
        self.current_session: Optional[BatchProgress] = None
        self.auto_save_interval = 10  # 秒
        self.last_save_time = 0
        
    def generate_session_id(self, target_path: str, mask: str, mode: str) -> str:
        """生成会话ID"""
        content = f"{target_path}_{mask}_{mode}_{datetime.now().strftime('%Y%m%d')}"
        return hashlib.md5(content.encode()).hexdigest()[:12]
    
    def create_session(self, target_path: str, mask: str, mode: str, 
                      file_list: List[str]) -> str:
        """创建新的破解会话"""
        session_id = self.generate_session_id(target_path, mask, mode)
        
        # 检查是否存在未完成的会话
        existing_session = self.load_session(session_id)
        if existing_session:
            console.print(f"[yellow]📄 发现未完成的会话: {session_id}[/yellow]")
            console.print(f"[cyan]   目标: {existing_session.target_path}[/cyan]")
            console.print(f"[cyan]   进度: {existing_session.completed_files}/{existing_session.total_files}[/cyan]")
            
            if self._confirm_resume():
                console.print("[green]✅ 恢复之前的会话[/green]")
                self.current_session = existing_session
                return session_id
            else:
                console.print("[yellow]🔄 创建新会话[/yellow]")
        
        # 创建新会话
        self.current_session = BatchProgress(
            session_id=session_id,
            target_path=target_path,
            mask=mask,
            mode=mode,
            total_files=len(file_list),
            completed_files=0,
            failed_files=0,
            skipped_files=0,
            start_time=datetime.now().isoformat(),
            last_update=datetime.now().isoformat()
        )
        
        # 初始化任务列表
        for file_path in file_list:
            task = TaskProgress(
                file_path=file_path,
                status="pending"
            )
            self.current_session.tasks.append(task)
        
        self.save_session()
        console.print(f"[green]✅ 创建新会话: {session_id}[/green]")
        return session_id
    
    def _confirm_resume(self) -> bool:
        """确认是否恢复会话"""
        try:
            from rich.prompt import Confirm
            return Confirm.ask("是否继续之前未完成的破解任务?")
        except:
            # 如果无法交互，默认恢复
            return True
    
    def get_pending_tasks(self) -> List[TaskProgress]:
        """获取待处理的任务"""
        if not self.current_session:
            return []
        
        return [task for task in self.current_session.tasks 
                if task.status == "pending"]
    
    def get_task_by_path(self, file_path: str) -> Optional[TaskProgress]:
        """根据文件路径获取任务"""
        if not self.current_session:
            return None
            
        for task in self.current_session.tasks:
            if task.file_path == file_path:
                return task
        return None
    
    def start_task(self, file_path: str) -> bool:
        """开始处理任务"""
        task = self.get_task_by_path(file_path)
        if not task:
            return False
            
        task.status = "processing"
        task.start_time = datetime.now().isoformat()
        task.attempts += 1
        
        self._auto_save()
        return True
    
    def complete_task(self, file_path: str, password: str, duration: float, 
                     alias: str = None, public_key_md5: str = None, public_key_sha1: str = None,
                     keystore_type: str = None, certificate_info: Dict[str, Any] = None) -> bool:
        """完成任务"""
        task = self.get_task_by_path(file_path)
        if not task:
            return False
            
        task.status = "completed"
        task.password = password
        task.end_time = datetime.now().isoformat()
        task.duration = duration
        task.alias = alias
        task.public_key_md5 = public_key_md5
        task.public_key_sha1 = public_key_sha1
        task.keystore_type = keystore_type
        task.certificate_info = certificate_info
        
        self.current_session.completed_files += 1
        self.current_session.last_update = datetime.now().isoformat()
        
        self._update_estimated_completion()
        self._auto_save()
        return True
    
    def fail_task(self, file_path: str, error_message: str) -> bool:
        """任务失败"""
        task = self.get_task_by_path(file_path)
        if not task:
            return False
            
        task.status = "failed"
        task.error_message = error_message
        task.end_time = datetime.now().isoformat()
        
        self.current_session.failed_files += 1
        self.current_session.last_update = datetime.now().isoformat()
        
        self._auto_save()
        return True
    
    def skip_task(self, file_path: str, reason: str) -> bool:
        """跳过任务"""
        task = self.get_task_by_path(file_path)
        if not task:
            return False
            
        task.status = "skipped"
        task.error_message = reason
        task.end_time = datetime.now().isoformat()
        
        self.current_session.skipped_files += 1
        self.current_session.last_update = datetime.now().isoformat()
        
        self._auto_save()
        return True
    
    def _update_estimated_completion(self):
        """更新预估完成时间"""
        if not self.current_session or self.current_session.completed_files == 0:
            return
            
        # 计算平均处理时间
        completed_tasks = [task for task in self.current_session.tasks 
                          if task.status == "completed" and task.duration]
        
        if not completed_tasks:
            return
            
        avg_duration = sum(task.duration for task in completed_tasks) / len(completed_tasks)
        remaining_files = (self.current_session.total_files - 
                          self.current_session.completed_files - 
                          self.current_session.failed_files - 
                          self.current_session.skipped_files)
        
        estimated_seconds = remaining_files * avg_duration
        estimated_time = datetime.now().timestamp() + estimated_seconds
        self.current_session.estimated_completion = datetime.fromtimestamp(estimated_time).isoformat()
    
    def _auto_save(self):
        """自动保存"""
        current_time = time.time()
        if current_time - self.last_save_time >= self.auto_save_interval:
            self.save_session()
            self.last_save_time = current_time
    
    def save_session(self):
        """保存会话"""
        if not self.current_session:
            return
            
        session_file = self.progress_dir / f"{self.current_session.session_id}.json"
        
        try:
            with open(session_file, 'w', encoding='utf-8') as f:
                json.dump(self.current_session.to_dict(), f, indent=2, ensure_ascii=False)
        except Exception as e:
            console.print(f"[red]❌ 保存会话失败: {e}[/red]")
    
    def load_session(self, session_id: str) -> Optional[BatchProgress]:
        """加载会话"""
        session_file = self.progress_dir / f"{session_id}.json"
        
        if not session_file.exists():
            return None
            
        try:
            with open(session_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return BatchProgress.from_dict(data)
        except Exception as e:
            console.print(f"[red]❌ 加载会话失败: {e}[/red]")
            return None
    
    def list_sessions(self) -> List[str]:
        """列出所有会话"""
        session_files = list(self.progress_dir.glob("*.json"))
        return [f.stem for f in session_files]
    
    def delete_session(self, session_id: str) -> bool:
        """删除会话"""
        session_file = self.progress_dir / f"{session_id}.json"
        
        try:
            if session_file.exists():
                session_file.unlink()
                return True
        except Exception as e:
            console.print(f"[red]❌ 删除会话失败: {e}[/red]")
        return False
    
    def show_progress(self):
        """显示当前进度"""
        if not self.current_session:
            console.print("[yellow]⚠️ 没有活动会话[/yellow]")
            return
            
        session = self.current_session
        
        # 计算进度百分比
        total_processed = session.completed_files + session.failed_files + session.skipped_files
        progress_percent = (total_processed / session.total_files * 100) if session.total_files > 0 else 0
        
        # 创建进度表格
        table = Table(title=f"📊 会话进度: {session.session_id}", border_style="blue")
        table.add_column("项目", style="cyan", width=15)
        table.add_column("值", style="white", width=30)
        
        table.add_row("目标路径", session.target_path)
        table.add_row("密码掩码", session.mask)
        table.add_row("破解模式", session.mode)
        table.add_row("总文件数", str(session.total_files))
        table.add_row("已完成", f"{session.completed_files} ✅")
        table.add_row("已失败", f"{session.failed_files} ❌")
        table.add_row("已跳过", f"{session.skipped_files} ⏭️")
        table.add_row("进度", f"{progress_percent:.1f}%")
        
        if session.estimated_completion:
            est_time = datetime.fromisoformat(session.estimated_completion)
            table.add_row("预计完成", est_time.strftime("%Y-%m-%d %H:%M:%S"))
        
        console.print(table)
    
    def get_results_summary(self) -> Dict[str, Any]:
        """获取结果汇总"""
        if not self.current_session:
            return {}
            
        successful_tasks = [task for task in self.current_session.tasks 
                           if task.status == "completed"]
        
        results = []
        for task in successful_tasks:
            file_path = Path(task.file_path)
            result_entry = {
                "路径": str(file_path.parent),
                "ID": file_path.parent.name,  # UUID文件夹名
                "文件名": file_path.name,
                "别名": task.alias or "未知",
                "私钥密码": task.password,
                "签名公钥MD5": task.public_key_md5 or "未提取",
                "签名公钥SHA1": task.public_key_sha1 or "未提取",
                "keystore类型": task.keystore_type or "JKS",
                "破解耗时": f"{task.duration:.2f}秒" if task.duration else "未知",
                "证书信息": task.certificate_info or {}
            }
            results.append(result_entry)
        
        return {
            "session_id": self.current_session.session_id,
            "破解时间": self.current_session.start_time,
            "目标路径": self.current_session.target_path,
            "密码掩码": self.current_session.mask,
            "破解模式": self.current_session.mode,
            "总文件数": self.current_session.total_files,
            "成功破解": len(successful_tasks),
            "破解失败": self.current_session.failed_files,
            "跳过文件": self.current_session.skipped_files,
            "成功率": f"{(len(successful_tasks) / max(self.current_session.total_files, 1) * 100):.1f}%",
            "破解结果": results
        }
    
    def export_results(self, output_file: Optional[str] = None, export_xlsx: bool = True) -> str:
        """导出结果到JSON和XLSX"""
        if not self.current_session:
            console.print("[red]❌ 没有活动会话可导出[/red]")
            return ""
            
        # 确定输出文件名
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = f"crack_results_{self.current_session.session_id}_{timestamp}"
            json_file = f"{base_name}.json"
            xlsx_file = f"{base_name}.xlsx"
        else:
            base_name = Path(output_file).stem
            json_file = f"{base_name}.json"
            xlsx_file = f"{base_name}.xlsx"
        
        results = self.get_results_summary()
        
        # 导出JSON文件
        try:
            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, ensure_ascii=False)
            console.print(f"[green]✅ JSON结果已导出到: {json_file}[/green]")
        except Exception as e:
            console.print(f"[red]❌ JSON导出失败: {e}[/red]")
            return ""
        
        # 导出XLSX文件
        if export_xlsx:
            try:
                xlsx_path = self._export_to_xlsx(results, xlsx_file)
                if xlsx_path:
                    console.print(f"[green]✅ Excel结果已导出到: {xlsx_path}[/green]")
                    return xlsx_path
            except Exception as e:
                console.print(f"[yellow]⚠️ Excel导出失败: {e}[/yellow]")
                console.print(f"[cyan]💡 提示: 请安装openpyxl: pip install openpyxl[/cyan]")
        
        # 如果没有导出Excel或Excel导出失败，但JSON成功，则给出提示
        if export_xlsx and not results.get("破解结果"):
            console.print(f"[yellow]ℹ️ 没有成功破解的结果，仅生成统计信息的Excel文件[/yellow]")
        
        return json_file
    
    def _export_to_xlsx(self, results: Dict[str, Any], xlsx_file: str) -> Optional[str]:
        """导出结果到Excel文件"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            console.print("[yellow]⚠️ 未安装openpyxl，跳过Excel导出[/yellow]")
            return None
        
        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        
        # 主结果表
        ws_main = wb.active
        ws_main.title = "破解结果"
        
        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 主表头
        headers = ["路径", "ID", "文件名", "别名", "私钥密码", "签名公钥MD5", "签名公钥SHA1", "keystore类型", "破解耗时"]
        for col, header in enumerate(headers, 1):
            cell = ws_main.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # 填充数据
        for row, result in enumerate(results.get("破解结果", []), 2):
            ws_main.cell(row=row, column=1, value=result.get("路径", "")).border = border
            ws_main.cell(row=row, column=2, value=result.get("ID", "")).border = border
            ws_main.cell(row=row, column=3, value=result.get("文件名", "")).border = border
            ws_main.cell(row=row, column=4, value=result.get("别名", "")).border = border
            ws_main.cell(row=row, column=5, value=result.get("私钥密码", "")).border = border
            ws_main.cell(row=row, column=6, value=result.get("签名公钥MD5", "")).border = border
            ws_main.cell(row=row, column=7, value=result.get("签名公钥SHA1", "")).border = border
            ws_main.cell(row=row, column=8, value=result.get("keystore类型", "")).border = border
            ws_main.cell(row=row, column=9, value=result.get("破解耗时", "")).border = border
        
        # 自动调整列宽
        for column in ws_main.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_main.column_dimensions[column_letter].width = adjusted_width
        
        # 统计信息表
        ws_stats = wb.create_sheet("统计信息")
        
        stats_data = [
            ["会话ID", results.get("session_id", "")],
            ["破解时间", results.get("破解时间", "")],
            ["目标路径", results.get("目标路径", "")],
            ["密码掩码", results.get("密码掩码", "")],
            ["破解模式", results.get("破解模式", "")],
            ["总文件数", results.get("总文件数", 0)],
            ["成功破解", results.get("成功破解", 0)],
            ["破解失败", results.get("破解失败", 0)],
            ["跳过文件", results.get("跳过文件", 0)],
            ["成功率", results.get("成功率", "0%")]
        ]
        
        for row, (key, value) in enumerate(stats_data, 1):
            ws_stats.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws_stats.cell(row=row, column=2, value=str(value))
            ws_stats.cell(row=row, column=1).border = border
            ws_stats.cell(row=row, column=2).border = border
        
        # 调整统计信息表列宽
        ws_stats.column_dimensions['A'].width = 15
        ws_stats.column_dimensions['B'].width = 30
        
        # 保存文件
        wb.save(xlsx_file)
        return xlsx_file
    
    def cleanup_completed_sessions(self, keep_days: int = 7):
        """清理已完成的旧会话"""
        current_time = time.time()
        cutoff_time = current_time - (keep_days * 24 * 60 * 60)
        
        cleaned = 0
        for session_file in self.progress_dir.glob("*.json"):
            try:
                # 检查文件修改时间
                if session_file.stat().st_mtime < cutoff_time:
                    session = self.load_session(session_file.stem)
                    if session:
                        # 只删除已完成的会话
                        total_processed = (session.completed_files + 
                                         session.failed_files + 
                                         session.skipped_files)
                        if total_processed >= session.total_files:
                            session_file.unlink()
                            cleaned += 1
            except Exception:
                continue
        
        if cleaned > 0:
            console.print(f"[green]🧹 清理了 {cleaned} 个已完成的旧会话[/green]") 